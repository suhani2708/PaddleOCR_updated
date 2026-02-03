"""
Comprehensive Excel exporter for PaddleOCR results.
Converts OCR results to Excel format with confidence-based highlighting and proper row/column organization.
"""

from typing import Dict, List, Tuple, Any, Optional, Literal
from pathlib import Path
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from paddleocr import PaddleOCR
import os
import cv2
import numpy as np

from text_row_detector import detect_text_rows_from_bounding_boxes
from confidence_colors import get_confidence_color
from ocr_utils import extract_word_data


def preprocess_image_for_ocr(image_path: str) -> str:
    """
    Preprocess an image to improve OCR accuracy.

    Args:
        image_path: Path to the input image

    Returns:
        Path to the preprocessed image (temporary file)
    """
    # Load the image
    img = cv2.imread(image_path)
    if img is None:
        raise ValueError(f"Could not load image: {image_path}")

    # Step 1: Convert to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # Step 2: Apply adaptive thresholding
    thresh = cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        11,
        2
    )

    # Step 3: Denoise the image
    denoised = cv2.fastNlMeansDenoising(thresh)

    # Step 4: Sharpen the text using a kernel filter
    kernel = np.array([[-1,-1,-1],
                       [-1, 9,-1],
                       [-1,-1,-1]])
    sharpened = cv2.filter2D(denoised, -1, kernel)

    # Step 5: Detect and correct skew/rotation
    coords = np.column_stack(np.where(sharpened > 0))
    angle = cv2.minAreaRect(coords)[-1]

    # Adjust angle based on convention
    if angle < -45:
        angle = -(90 + angle)
    else:
        angle = -angle

    # Rotate the image to correct skew
    (h, w) = sharpened.shape[:2]
    center = (w // 2, h // 2)
    M = cv2.getRotationMatrix2D(center, angle, 1.0)
    rotated = cv2.warpAffine(sharpened, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)

    # Save the preprocessed image to a temporary file
    temp_path = image_path.replace('.jpg', '_preprocessed.jpg').replace('.png', '_preprocessed.png').replace('.jpeg', '_preprocessed.jpeg')
    cv2.imwrite(temp_path, rotated)

    return temp_path


class PaddleOCRExcelExporter:
    """
    A comprehensive class to export PaddleOCR results to Excel format with confidence-based highlighting.
    
    This class provides a complete workflow for processing images with OCR, organizing
    the results into rows, and exporting them to Excel with visual indicators for
    low-confidence text that may require human verification.
    
    Attributes:
        ocr (PaddleOCR): The PaddleOCR instance for performing OCR
        confidence_threshold (float): Threshold below which text is considered low confidence
        enable_orientation (bool): Whether to enable text orientation detection
    """
    
    def __init__(self, lang: str = 'en', confidence_threshold: float = 1.00, 
                 enable_orientation: bool = True, return_word_box: bool = True,
                 y_tolerance: int = 10, reading_direction: Literal['ltr', 'rtl', 'auto'] = 'ltr'):
        """
        Initialize the PaddleOCRExcelExporter.
        
        Args:
            lang: Language for OCR processing (default: 'en')
            confidence_threshold: Threshold below which text is highlighted (default: 1.00)
            enable_orientation: Whether to enable text orientation detection (default: True)
            return_word_box: Whether to return word-level bounding boxes (default: True)
            y_tolerance: Tolerance for row detection based on Y-coordinates (default: 10)
            reading_direction: Reading direction for text ('ltr', 'rtl', 'auto') (default: 'ltr')
        """
        # Input validation
        if not isinstance(lang, str):
            raise TypeError("Language must be a string")
        
        if not (0.0 <= confidence_threshold <= 1.0):
            raise ValueError(f"Confidence threshold must be between 0.0 and 1.0, got {confidence_threshold}")
        
        if not isinstance(enable_orientation, bool):
            raise TypeError("enable_orientation must be a boolean")
        
        if not isinstance(return_word_box, bool):
            raise TypeError("return_word_box must be a boolean")
        
        if not isinstance(y_tolerance, int) or y_tolerance < 0:
            raise ValueError(f"Y tolerance must be a non-negative integer, got {y_tolerance}")
        
        if reading_direction not in ['ltr', 'rtl', 'auto']:
            raise ValueError(f"Reading direction must be 'ltr', 'rtl', or 'auto', got {reading_direction}")
        
        # Store configuration
        self.lang = lang
        self.confidence_threshold = confidence_threshold
        self.enable_orientation = enable_orientation
        self.return_word_box = return_word_box
        self.y_tolerance = y_tolerance
        self.reading_direction = reading_direction
        
        try:
            # Initialize PaddleOCR with specified parameters
            # use_angle_cls enables text line orientation classification (rotated text detection)
            self.ocr = PaddleOCR(
                lang=lang,
                use_angle_cls=enable_orientation,
                return_word_box=return_word_box
            )
        except Exception as e:
            logging.error(f"Failed to initialize PaddleOCR: {str(e)}")
            raise
        
        # Set up logging
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
    
    def set_confidence_threshold(self, threshold: float) -> None:
        """
        Set the confidence threshold for highlighting low-confidence text.
        
        Args:
            threshold: New confidence threshold value (between 0.0 and 1.0)
        """
        if not (0.0 <= threshold <= 1.0):
            raise ValueError(f"Confidence threshold must be between 0.0 and 1.0, got {threshold}")
        
        self.confidence_threshold = threshold
        self.logger.info(f"Confidence threshold updated to {threshold}")

    def set_y_tolerance(self, y_tolerance: int) -> None:
        """
        Set the Y-coordinate tolerance for row detection.
        
        Args:
            y_tolerance: New Y-coordinate tolerance value (non-negative integer)
        """
        if not isinstance(y_tolerance, int) or y_tolerance < 0:
            raise ValueError(f"Y tolerance must be a non-negative integer, got {y_tolerance}")
        
        self.y_tolerance = y_tolerance
        self.logger.info(f"Y tolerance updated to {y_tolerance}")

    def set_reading_direction(self, reading_direction: str) -> None:
        """
        Set the reading direction for text processing.
        
        Args:
            reading_direction: New reading direction ('ltr', 'rtl', or 'auto')
        """
        if reading_direction not in ['ltr', 'rtl', 'auto']:
            raise ValueError(f"Reading direction must be 'ltr', 'rtl', or 'auto', got {reading_direction}")
        
        self.reading_direction = reading_direction
        self.logger.info(f"Reading direction updated to {reading_direction}")

    def get_config(self) -> Dict[str, Any]:
        """
        Get the current configuration settings.
        
        Returns:
            Dictionary containing all configuration parameters
        """
        return {
            'lang': self.lang,
            'confidence_threshold': self.confidence_threshold,
            'enable_orientation': self.enable_orientation,
            'return_word_box': self.return_word_box,
            'y_tolerance': self.y_tolerance,
            'reading_direction': self.reading_direction
        }
    
    def preprocess_image_for_orientation(self, image_path: str) -> str:
        """
        Preprocess an image to handle rotation and improve OCR results.

        Args:
            image_path: Path to the input image

        Returns:
            Path to the preprocessed image (may be the same as input if no preprocessing needed)
        """
        try:
            # Use the comprehensive preprocessing function
            preprocessed_path = preprocess_image_for_ocr(image_path)
            return preprocessed_path

        except Exception as e:
            self.logger.warning(f"Image preprocessing failed: {str(e)}, using original image")
            return image_path

    def process_image(self, image_path: str) -> List[Dict[str, Any]]:
        """
        Process an image with OCR and extract structured word data.
        
        Args:
            image_path: Path to the image file to process
            
        Returns:
            List of dictionaries containing word data with text, coordinates, and confidence
        """
        self.logger.info(f"Processing image: {image_path}")
        
        # Validate image path
        if not isinstance(image_path, str):
            raise TypeError("Image path must be a string")
        
        if not os.path.exists(image_path):
            raise FileNotFoundError(f"Image not found: {image_path}")
        
        # Check if image is readable
        if not os.access(image_path, os.R_OK):
            raise PermissionError(f"Cannot read image file: {image_path}")
        
        # Validate image format
        try:
            image_extension = Path(image_path).suffix.lower()
            valid_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.gif']
            if image_extension not in valid_extensions:
                raise ValueError(f"Unsupported image format: {image_extension}. Supported formats: {valid_extensions}")
            
            # Try to load the image with OpenCV to verify it's a valid image
            img = cv2.imread(image_path)
            if img is None:
                raise ValueError(f"Cannot read image file: {image_path}. File may be corrupted or in unsupported format.")
        except Exception as e:
            self.logger.error(f"Image format validation failed: {str(e)}")
            raise
        
        try:
            # Preprocess image for better orientation handling
            processed_image_path = self.preprocess_image_for_orientation(image_path)
            
            # Run OCR on image with text orientation detection enabled
            result = self.ocr.ocr(
                processed_image_path, 
                cls=self.enable_orientation
            )
            
            # Handle OCR failures (empty results)
            if not result:
                self.logger.warning(f"No text detected in image: {image_path}")
                return []
            
            # Extract word-level data
            word_data = extract_word_data(result)
            
            if not word_data:
                self.logger.warning(f"No word data extracted from image: {image_path}")
                return []
            
            self.logger.info(f"Extracted {len(word_data)} words from image")
            
            # Organize into rows based on Y-coordinates using configurable parameters
            detections = [(item['x'], item['text'], item['confidence'], item['y']) for item in word_data]
            rows = detect_text_rows_from_bounding_boxes(detections, y_tolerance=self.y_tolerance, reading_direction=self.reading_direction)
            
            # Convert to list of rows, each containing a list of words
            rows_list = []
            for y_coord, words in sorted(rows.items()):
                row_data = []
                for x_pos, text, confidence in words:
                    # Validate confidence values
                    if not (0.0 <= confidence <= 1.0):
                        self.logger.warning(f"Invalid confidence value: {confidence}. Setting to 0.0")
                        confidence = 0.0
                    
                    row_data.append({
                        'text': text,
                        'confidence': confidence,
                        'x': x_pos,
                        'y': y_coord
                    })
                rows_list.append(row_data)
            
            self.logger.info(f"Organized words into {len(rows_list)} rows")
            
            return rows_list
            
        except Exception as e:
            self.logger.error(f"Error processing image {image_path}: {str(e)}")
            raise
    
    def create_excel_workbook(self, rows_data: List[List[Dict[str, Any]]]) -> Workbook:
        """
        Create an Excel workbook from processed OCR data.

        Args:
            rows_data: List of rows, each containing a list of word dictionaries
                      Each word dictionary has 'text', 'confidence', 'x', 'y' keys

        Returns:
            openpyxl.Workbook object with formatted data
        """
        self.logger.info("Creating Excel workbook")

        # Validate input data
        if not isinstance(rows_data, list):
            raise TypeError("rows_data must be a list")

        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "OCR Results"

            # Add header row explaining the color scheme
            header = ["Word 1", "Word 2", "Word 3", "Word 4", "..."]  # Dynamic based on content
            ws.append(header)

            # Add color scheme explanation in a separate row
            color_explanation = ["Color Guide:", "Green = High Confidence", "Yellow = Medium Confidence", "Red = Low Confidence", ""]
            ws.append(color_explanation)

            # Add an empty row before the actual data
            ws.append([])

            # Process each row of detected text
            for row_data in rows_data:
                if not isinstance(row_data, list):
                    self.logger.warning(f"Skipping invalid row data: {row_data}")
                    continue

                # Sort the row data by x-coordinate to maintain left-to-right order
                sorted_row_data = sorted(row_data, key=lambda item: item.get('x', 0))

                # Create a new Excel row with each word in a separate column
                row_values = []
                for word_info in sorted_row_data:
                    if not isinstance(word_info, dict) or 'text' not in word_info:
                        self.logger.warning(f"Skipping invalid word data: {word_info}")
                        continue

                    text = word_info['text']
                    row_values.append(text)

                # Add the row to the worksheet
                if row_values:  # Only add if there's data
                    ws.append(row_values)

                    # Get the current row index
                    current_row_idx = ws.max_row

                    # Apply background color based on confidence using get_confidence_color function
                    for col_idx, word_info in enumerate(sorted_row_data):
                        if 'confidence' not in word_info:
                            continue  # Skip if no confidence value

                        confidence = word_info['confidence']

                        # Validate confidence value
                        if not isinstance(confidence, (int, float)) or not (0.0 <= confidence <= 1.0):
                            self.logger.warning(f"Invalid confidence value: {confidence}. Using 0.0")
                            confidence = 0.0

                        try:
                            # Get the appropriate color based on confidence
                            hex_color = get_confidence_color(confidence)

                            # Create a fill with the appropriate gradient color
                            cell_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

                            # Apply the fill to the cell
                            cell = ws.cell(row=current_row_idx, column=col_idx + 1)  # +1 because Excel columns start at 1
                            cell.fill = cell_fill
                        except Exception as color_error:
                            self.logger.warning(f"Could not apply color to cell: {color_error}")
                            # Continue without coloring if color application fails

            # Auto-adjust column widths for readability
            try:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    # Adjust width based on content, with reasonable bounds
                    adjusted_width = min(max(max_length + 2, 10), 50)  # Min 10, Max 50, based on content
                    ws.column_dimensions[column_letter].width = adjusted_width
            except Exception as width_error:
                self.logger.warning(f"Could not adjust column widths: {width_error}")

        except Exception as e:
            self.logger.error(f"Error creating Excel workbook: {str(e)}")
            raise

        self.logger.info("Excel workbook created successfully with confidence-based coloring")
        return wb
    
    def export_to_excel(self, image_path: str, excel_path: str) -> str:
        """
        Complete workflow: Process image with OCR and export results to Excel.
        
        Args:
            image_path: Path to the input image file
            excel_path: Path for the output Excel file
            
        Returns:
            Path to the created Excel file
        """
        self.logger.info(f"Starting export process: {image_path} -> {excel_path}")
        
        # Validate input parameters
        if not isinstance(image_path, str):
            raise TypeError("Image path must be a string")
        
        if not isinstance(excel_path, str):
            raise TypeError("Excel path must be a string")
        
        # Validate output path
        excel_path_obj = Path(excel_path)
        output_dir = excel_path_obj.parent
        
        # Check if output directory is writable
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            # Create a temporary file to test write permissions
            test_file = output_dir / ".write_test"
            test_file.touch()
            test_file.unlink()  # Remove the test file
        except Exception as e:
            raise PermissionError(f"Cannot write to output directory: {output_dir}. Error: {str(e)}")
        
        # Check if output file extension is valid
        if excel_path_obj.suffix.lower() not in ['.xlsx', '.xls']:
            raise ValueError(f"Invalid Excel file extension: {excel_path_obj.suffix}. Use .xlsx or .xls")
        
        try:
            # Process the image
            rows_data = self.process_image(image_path)
            
            # Create Excel workbook
            wb = self.create_excel_workbook(rows_data)
            
            # Save workbook with error handling
            try:
                wb.save(excel_path)
            except PermissionError:
                raise PermissionError(f"Permission denied when saving Excel file: {excel_path}")
            except OSError as e:
                if "Disk full" in str(e) or e.errno == 28:  # errno 28 = No space left on device
                    raise OSError(f"Not enough disk space to save Excel file: {excel_path}")
                else:
                    raise OSError(f"Error saving Excel file: {str(e)}")
            except Exception as e:
                raise Exception(f"Unexpected error saving Excel file: {str(e)}")
            
            self.logger.info(f"Export completed successfully: {excel_path}")
            return excel_path

        except FileNotFoundError:
            self.logger.error(f"Image file not found: {image_path}")
            raise
        except PermissionError:
            self.logger.error(f"Permission error during export: {image_path} -> {excel_path}")
            raise
        except OSError as e:
            self.logger.error(f"OS error during export: {str(e)}")
            raise
        except Exception as e:
            self.logger.error(f"Error during export process: {str(e)}")
            raise

    def get_quality_metrics(self, image_path: str) -> Dict[str, Any]:
        """
        Calculate quality metrics for the OCR processing of an image.

        Args:
            image_path: Path to the image file to analyze

        Returns:
            Dictionary containing quality metrics
        """
        try:
            # Process the image to get word data
            rows_data = self.process_image(image_path)

            if not rows_data:
                return {
                    'total_rows': 0,
                    'total_words': 0,
                    'avg_confidence': 0.0,
                    'min_confidence': 0.0,
                    'max_confidence': 0.0,
                    'low_confidence_words': 0,
                    'low_confidence_percentage': 0.0,
                    'processing_time': 0.0,
                    'confidence_distribution': {}
                }

            all_confidences = []
            total_words = 0
            low_conf_words = 0

            for row in rows_data:
                for word_info in row:
                    if 'confidence' in word_info:
                        conf = word_info['confidence']
                        all_confidences.append(conf)
                        total_words += 1
                        if conf < self.confidence_threshold:
                            low_conf_words += 1

            if not all_confidences:
                return {
                    'total_rows': len(rows_data),
                    'total_words': 0,
                    'avg_confidence': 0.0,
                    'min_confidence': 0.0,
                    'max_confidence': 0.0,
                    'low_confidence_words': 0,
                    'low_confidence_percentage': 0.0,
                    'processing_time': 0.0,
                    'confidence_distribution': {}
                }

            avg_conf = sum(all_confidences) / len(all_confidences)
            min_conf = min(all_confidences)
            max_conf = max(all_confidences)
            low_conf_percent = (low_conf_words / total_words) * 100 if total_words > 0 else 0

            # Calculate confidence distribution
            confidence_ranges = {
                'very_high': len([c for c in all_confidences if c >= 0.9]),
                'high': len([c for c in all_confidences if 0.7 <= c < 0.9]),
                'medium': len([c for c in all_confidences if 0.5 <= c < 0.7]),
                'low': len([c for c in all_confidences if 0.3 <= c < 0.5]),
                'very_low': len([c for c in all_confidences if c < 0.3])
            }

            return {
                'total_rows': len(rows_data),
                'total_words': total_words,
                'avg_confidence': avg_conf,
                'min_confidence': min_conf,
                'max_confidence': max_conf,
                'low_confidence_words': low_conf_words,
                'low_confidence_percentage': low_conf_percent,
                'confidence_distribution': confidence_ranges
            }

        except Exception as e:
            self.logger.error(f"Error calculating quality metrics: {str(e)}")
            raise

    def generate_report(self, image_path: str) -> str:
        """
        Generate a comprehensive report about the OCR processing of an image.

        Args:
            image_path: Path to the image file to analyze

        Returns:
            String containing the formatted report
        """
        metrics = self.get_quality_metrics(image_path)

        report = f"""
OCR Processing Report
===================

File: {image_path}
Date: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Statistics:
- Total rows detected: {metrics['total_rows']}
- Total words extracted: {metrics['total_words']}
- Average confidence: {metrics['avg_confidence']:.2f}
- Min confidence: {metrics['min_confidence']:.2f}
- Max confidence: {metrics['max_confidence']:.2f}

Quality Assessment:
- Low confidence words: {metrics['low_confidence_words']} ({metrics['low_confidence_percentage']:.1f}%)
- Very high confidence (≥0.9): {metrics['confidence_distribution']['very_high']}
- High confidence (0.7-0.9): {metrics['confidence_distribution']['high']}
- Medium confidence (0.5-0.7): {metrics['confidence_distribution']['medium']}
- Low confidence (0.3-0.5): {metrics['confidence_distribution']['low']}
- Very low confidence (<0.3): {metrics['confidence_distribution']['very_low']}

Recommendation:
"""
        if metrics['low_confidence_percentage'] > 30:
            report += "- High percentage of low-confidence text detected. Manual review recommended.\n"
        elif metrics['low_confidence_percentage'] > 10:
            report += "- Moderate percentage of low-confidence text detected. Spot check recommended.\n"
        else:
            report += "- Good quality OCR results. Minimal manual review needed.\n"

        return report

    def create_custom_template(self, template_name: str, config: Dict[str, Any]) -> None:
        """
        Create a custom output template for different document types.

        Args:
            template_name: Name of the template
            config: Configuration for the template
        """
        # Store template configuration
        if not hasattr(self, '_templates'):
            self._templates = {}

        self._templates[template_name] = config
        self.logger.info(f"Created custom template: {template_name}")

    def export_with_template(self, image_path: str, excel_path: str, template_name: str) -> str:
        """
        Export OCR results using a custom template.

        Args:
            image_path: Path to the input image file
            excel_path: Path for the output Excel file
            template_name: Name of the template to use

        Returns:
            Path to the created Excel file
        """
        if not hasattr(self, '_templates') or template_name not in self._templates:
            raise ValueError(f"Template '{template_name}' not found")

        template_config = self._templates[template_name]

        # Process the image
        rows_data = self.process_image(image_path)

        # Create workbook with template-specific formatting
        wb = self._apply_template_formatting(rows_data, template_config)

        # Ensure output directory exists
        output_dir = Path(excel_path).parent
        output_dir.mkdir(parents=True, exist_ok=True)

        # Save workbook
        wb.save(excel_path)

        self.logger.info(f"Export completed with template '{template_name}': {excel_path}")
        return excel_path

    def _apply_template_formatting(self, rows_data: List[List[Dict[str, Any]]],
                                  template_config: Dict[str, Any]) -> Workbook:
        """
        Apply template-specific formatting to the Excel workbook.

        Args:
            rows_data: Processed OCR data
            template_config: Template configuration

        Returns:
            Formatted Excel workbook
        """
        wb = Workbook()
        ws = wb.active
        ws.title = template_config.get('sheet_title', 'OCR Results')

        # Apply template-specific header
        headers = template_config.get('headers', ['Text', 'Confidence', 'Position_X', 'Position_Y'])
        ws.append(headers)

        # Add template-specific styling
        if template_config.get('include_color_guide', True):
            color_explanation = template_config.get('color_guide', [
                "Color Guide:", "Green = High Confidence",
                "Yellow = Medium Confidence", "Red = Low Confidence", ""
            ])
            ws.append(color_explanation)

        # Add an empty row before the actual data
        ws.append([])

        # Process each row of detected text according to template
        for row_data in rows_data:
            if not isinstance(row_data, list):
                continue

            # Format row according to template
            row_values = []
            for word_info in row_data:
                if not isinstance(word_info, dict) or 'text' not in word_info:
                    continue

                # Apply template-specific field extraction
                fields = template_config.get('fields', ['text'])
                for field in fields:
                    if field == 'text':
                        row_values.append(word_info.get('text', ''))
                    elif field == 'confidence':
                        row_values.append(f"{word_info.get('confidence', 0):.2f}")
                    elif field == 'x':
                        row_values.append(str(word_info.get('x', 0)))
                    elif field == 'y':
                        row_values.append(str(word_info.get('y', 0)))
                    else:
                        row_values.append(str(word_info.get(field, '')))

            # Add the row to the worksheet
            if row_values:
                ws.append(row_values)

                # Apply template-specific coloring
                current_row_idx = ws.max_row
                if template_config.get('apply_confidence_coloring', True):
                    for col_idx, word_info in enumerate(row_data):
                        if 'confidence' in word_info:
                            confidence = word_info['confidence']
                            hex_color = get_confidence_color(confidence)
                            cell_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

                            # Apply to the appropriate cells based on template
                            num_fields = len(template_config.get('fields', ['text']))
                            for offset in range(num_fields):
                                cell = ws.cell(row=current_row_idx, column=col_idx * num_fields + offset + 1)
                                cell.fill = cell_fill

        # Apply template-specific column width adjustments
        if template_config.get('auto_adjust_columns', True):
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        return wb


# Example usage of different templates
DEFAULT_TEMPLATES = {
    'standard': {
        'sheet_title': 'OCR Results',
        'headers': ['Text', 'Confidence', 'Position_X', 'Position_Y'],
        'fields': ['text', 'confidence', 'x', 'y'],
        'include_color_guide': True,
        'apply_confidence_coloring': True,
        'auto_adjust_columns': True
    },
    'simple': {
        'sheet_title': 'Simple OCR Output',
        'headers': ['Text', 'Confidence'],
        'fields': ['text', 'confidence'],
        'include_color_guide': False,
        'apply_confidence_coloring': True,
        'auto_adjust_columns': True
    },
    'detailed': {
        'sheet_title': 'Detailed OCR Analysis',
        'headers': ['Text', 'Confidence', 'X_Coord', 'Y_Coord', 'Element_Type'],
        'fields': ['text', 'confidence', 'x', 'y'],
        'include_color_guide': True,
        'apply_confidence_coloring': True,
        'auto_adjust_columns': True
    }
}


def example_usage():
    """
    Example usage of the preprocessing function and PaddleOCRExcelExporter.
    """
    # Example of using the preprocessing function directly
    print("Example: Preprocessing an image for OCR")
    print("Input: image_path = 'example.jpg'")
    print("Output: preprocessed image path")

    # Example of using the PaddleOCRExcelExporter with preprocessing
    print("\nExample: Using PaddleOCRExcelExporter with preprocessing")
    exporter = PaddleOCRExcelExporter(
        lang='en',
        confidence_threshold=0.8,
        enable_orientation=True,
        y_tolerance=10,
        reading_direction='ltr'
    )

    # The exporter will automatically use the preprocessing function
    # when processing images
    print("PaddleOCRExcelExporter initialized with preprocessing enabled")
    print("When calling process_image() or export_to_excel(), images will be preprocessed")
    print("to improve OCR accuracy through:")
    print("  - Conversion to grayscale")
    print("  - Adaptive thresholding")
    print("  - Denoising")
    print("  - Text sharpening")
    print("  - Skew/rotation correction")


if __name__ == "__main__":
    example_usage()