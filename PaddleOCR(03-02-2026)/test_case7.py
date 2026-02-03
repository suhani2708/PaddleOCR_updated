# -*- coding: utf-8 -*-
import os
import re
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def process_mm01s_datasets(image_path, output_excel="mm01s_output.xlsx"):
    """
    Process datasets based on MM01S pattern detection
    Args:
        image_path (str): Path to input image file
        output_excel (str): Output Excel filename
    """
    # Normalize file path to handle backslashes safely
    image_path = os.path.normpath(image_path)
    if not os.path.exists(image_path):
        print(f" Error: File not found - {image_path}")
        return

    # Initialize PaddleOCR with safe parameters
    print(" Starting OCR processing...")
    ocr = PaddleOCR(
        use_angle_cls=True,
        lang='en',
        use_gpu=False,
        show_log=False  # This is valid in recent PaddleOCR versions
    )

    # Run OCR on image
    result = ocr.ocr(image_path, cls=True)  # Correct API usage: using ocr() method
    if not result:
        print(" No text detected by OCR")
        return

    # Extract all detections with bounding box info
    all_detections = []
    # Handle PaddleOCR output format: result[0] contains all detected text regions
    # Each region is: [bounding_box, (text, confidence)]
    # bounding_box is [[x1,y1], [x2,y2], [x3,y3], [x4,y4]]
    # text is a string, confidence is a float
    items = result[0] if len(result) == 1 and isinstance(result[0], list) else result
    for line in items:
        if len(line) < 2:
            continue
        bbox = line[0]  # List of 4 points: [[x1,y1], [x2,y2], [x3,y3], [x4,y4]]
        text_info = line[1]
        if isinstance(text_info, (list, tuple)) and len(text_info) >= 2:
            text = text_info[0]
            confidence = text_info[1]
            x_start = bbox[0][0]
            y_top = min(pt[1] for pt in bbox)  # Topmost Y coordinate
            all_detections.append((text.strip(), confidence, x_start, y_top))

    # Sort by top Y (vertical position), then by X (horizontal position)
    all_detections.sort(key=lambda x: (x[3], x[2]))

    # Now build lines_with_conf as before (without y_top)
    lines_with_conf = [(text, conf, x) for text, conf, x, y in all_detections]

    # Group lines into datasets using MM01S pattern as delimiter
    datasets = []
    current_dataset = []
    for text, conf, x_pos in lines_with_conf:
        if re.match(r'^MM01S\d+', text):
            if current_dataset:
                datasets.append(current_dataset)
            current_dataset = [(text, conf)]
        else:
            if current_dataset:
                current_dataset.append((text, conf))
    
    if current_dataset:
        datasets.append(current_dataset)

    print(f" Total datasets detected: {len(datasets)}")

    if not datasets:
        print(" No valid MM01S datasets found")
        return

    # Convert each dataset to list of words (split on any whitespace)
    dataset_rows = []
    low_conf_rows = []

    for ds_idx, dataset in enumerate(datasets):
        full_text = " ".join([text for text, _ in dataset])
        words = full_text.split()  # Split on any whitespace → each word in separate column
        dataset_rows.append(words)

        avg_conf = sum(conf for _, conf in dataset) / len(dataset)
        if avg_conf < 0.99:
            low_conf_rows.append(ds_idx)

    # Normalize rows to same column count
    max_cols = max(len(row) for row in dataset_rows) if dataset_rows else 0
    normalized_rows = [
        row + [''] * (max_cols - len(row)) for row in dataset_rows
    ]

    # Create Excel without headers
    wb = Workbook()
    ws = wb.active
    ws.title = "MM01S Datasets"

    for row_data in normalized_rows:
        ws.append(row_data)

    # Highlight low-confidence rows in red
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    bold_red = Font(color="990000", bold=True)

    for row_idx in low_conf_rows:
        for col in range(1, max_cols + 1):
            cell = ws.cell(row=row_idx + 1, column=col)
            cell.fill = red_fill
            cell.font = bold_red

    # Save
    wb.save(output_excel)
    print(f" Successfully saved: {output_excel}")
    print(f" Data structure: {len(normalized_rows)} rows x {max_cols} columns")


    print("\nAll extracted datasets:")
    for i, row in enumerate(normalized_rows):
       print(f"Row {i+1}: {row}")


def validate_and_normalize_path(image_path):
    """Validate and normalize the image path"""
    if not image_path:
        raise ValueError("Image path cannot be empty")

    # Convert to absolute path if relative
    if not os.path.isabs(image_path):
        image_path = os.path.join(os.path.dirname(__file__), image_path)

    # Check if file exists
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"Image not found: {image_path}")

    return image_path

# Main execution
if __name__ == "__main__":
    IMAGE_PATH = 'docs/images/Statista-FormF-RKJJ-NA-1929751_d.jpg'  # Changed to relative path

    try:
        IMAGE_PATH = validate_and_normalize_path(IMAGE_PATH)
        process_mm01s_datasets(IMAGE_PATH, output_excel="mm01s_output.xlsx")
    except FileNotFoundError as e:
        print(f"Error: {e}")
    except ValueError as e:
        print(f"Error: {e}")