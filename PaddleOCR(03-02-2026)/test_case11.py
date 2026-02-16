# -*- coding: utf-8 -*-
import os
import re
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Configuration
IMAGE_PATH = 'docs/images/Statista-FormF-RKJJ-NA-1929751_d.jpg'
OUTPUT_FILE = 'updated_table.xlsx'
Y_TOLERANCE = 35  # Increased for slanted text
MM_PATTERN = r'^MM[\d\*#\-_]+$'

def is_mm_starter(text):
    """Check if text matches MM pattern (case-insensitive, flexible matching)"""
    text_clean = text.strip().upper()
    return bool(re.match(r'^MM[\d\*#\-_]{3,}$', text_clean)) or (text_clean.startswith('MM') and len(text_clean) >= 5)

def extract_and_combine_records(image_path, output_excel, y_tol=35):
    """Extract text and combine 4-5 lines per MM-started record into single Excel row"""
    
    # Check if image exists
    if not os.path.exists(image_path):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        image_path = os.path.join(script_dir, os.path.basename(image_path))
        if not os.path.exists(image_path):
            print(f"\n❌ ERROR: Image not found!")
            print(f"Searched for: {os.path.basename(image_path)}")
            print(f"In folder: {script_dir}")
            return None
    
    # print(f"\n✓ Processing: {os.path.basename(image_path)}")
    
    # Initialize OCR
    # print("[1/4] Initializing OCR...")
    ocr = PaddleOCR(lang='en', use_gpu=False, show_log=False)
    # print("✓ Done")
    
    # Extract text + coordinates
    # print("[2/4] Extracting text...")
    result = ocr.ocr(image_path, cls=True)
    
    if not result or not result[0]:
        print("❌ No text detected!")
        return None
    
    # print(f"✓ Found {len(result[0])} text regions")
    
    # Prepare boxes with centers
    boxes = []
    for det in result[0]:
        coords = det[0]
        text = det[1][0].strip()
        if not text or len(text) < 2:
            continue
        conf = det[1][1] * 100
        
        y_vals = [p[1] for p in coords]
        x_vals = [p[0] for p in coords]
        
        boxes.append({
            'text': text,
            'conf': conf,
            'x': (min(x_vals) + max(x_vals)) / 2,
            'y': (min(y_vals) + max(y_vals)) / 2,
        })
    
    if not boxes:
        print("❌ No valid text after filtering")
        return None
    
    # Sort by Y then X
    boxes.sort(key=lambda b: (b['y'], b['x']))
    
    # Group into visual lines using Y tolerance
    lines = []
    current_line = [boxes[0]]
    for box in boxes[1:]:
        if abs(box['y'] - current_line[0]['y']) <= y_tol:
            current_line.append(box)
        else:
            current_line.sort(key=lambda b: b['x'])
            lines.append(current_line)
            current_line = [box]
    current_line.sort(key=lambda b: b['x'])
    lines.append(current_line)
    
    # print(f"✓ Grouped into {len(lines)} visual lines (Y tolerance: {y_tol}px)")
    
    # Identify records and combine 4-5 lines per MM-started record
    records = []
    i = 0
    while i < len(lines):
        line = lines[i]
        is_record_start = any(is_mm_starter(box['text']) for box in line[:2])
        
        if is_record_start:
            record_lines = []
            lines_collected = 0
            max_lines = 5
            
            while lines_collected < max_lines and (i + lines_collected) < len(lines):
                current_idx = i + lines_collected
                check_line = lines[current_idx]
                
                if lines_collected > 0:
                    next_is_mm = any(is_mm_starter(box['text']) for box in check_line[:2])
                    if next_is_mm:
                        break
                
                record_lines.append(check_line)
                lines_collected += 1
            
            record_words = []
            record_confs = []
            for rec_line in record_lines:
                for box in rec_line:
                    record_words.append(box['text'])
                    record_confs.append(box['conf'])
            
            if record_words:
                records.append({'words': record_words, 'confs': record_confs})
            
            i += lines_collected
        else:
            i += 1
    
    # print(f"✓ Identified {len(records)} logical records (4-5 lines each)")
    
    if not records:
        print("⚠️  No MM-started records found. Using fallback: grouping every 5 lines as record.")
        for i in range(0, len(lines), 5):
            record_words = []
            record_confs = []
            for line in lines[i:i+5]:
                for box in line:
                    record_words.append(box['text'])
                    record_confs.append(box['conf'])
            if record_words:
                records.append({'words': record_words, 'confs': record_confs})
        # print(f"✓ Created {len(records)} fallback records")
    
    # Save to Excel
    # print("[3/4] Saving to Excel...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Records"
    
    low_conf_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    for row_idx, record in enumerate(records):
        for col_idx, (word, conf) in enumerate(zip(record['words'], record['confs'])):
            cell = ws.cell(row_idx + 1, col_idx + 1, word)
            if conf < 100:
                cell.fill = low_conf_fill
    
    max_cols = min(max(len(r['words']) for r in records), 70) if records else 10
    for col_idx in range(max_cols):
        col_letter = get_column_letter(col_idx + 1)
        ws.column_dimensions[col_letter].width = 18
    
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), output_excel)
    wb.save(output_path)
    
    # print(f"✓ Saved: {output_path}")
    # print(f"\n{'='*80}")
    # print(" SUCCESS - RECORD COMBINATION COMPLETE!")
    # print("="*80)
    # print(f" Input image : {os.path.basename(image_path)}")
    # print(f" Output file : {output_excel}")  # ✅ FIXED: changed from output_file to output_excel
    # print(f" Records     : {len(records)} (each 4-5 image lines → 1 Excel row)")
    # print(f" Avg words   : {sum(len(r['words']) for r in records) // len(records) if records else 0}")
    # print(f" MM pattern  : {MM_PATTERN}")
    # print(f" Y tolerance : {y_tol}px (handles slanted text)")
    # print("="*80)
    
    return output_path

def process_batch(image_folder, output_folder='output_combined'):
    """Process all images in folder (supports your 8 images)"""
    os.makedirs(output_folder, exist_ok=True)
    
    exts = ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.webp']
    image_files = [
        os.path.join(image_folder, f) 
        for f in os.listdir(image_folder) 
        if os.path.splitext(f)[1].lower() in exts
    ]
    
    if not image_files:
        print(f"❌ No images found in {image_folder}")
        return
    
    print(f"\n✓ Found {len(image_files)} images to process")
    
    for idx, img_path in enumerate(sorted(image_files), 1):
        print(f"\n{'='*80}")
        print(f"IMAGE {idx}/{len(image_files)}: {os.path.basename(img_path)}")
        print('='*80)
        
        output_name = f"combined_{idx}_{os.path.splitext(os.path.basename(img_path))[0]}.xlsx"
        output_path = os.path.join(output_folder, output_name)
        extract_and_combine_records(img_path, output_path, Y_TOLERANCE)
    
    print(f"\n✓ All {len(image_files)} images processed!")
    print(f"  Outputs saved to: {os.path.abspath(output_folder)}")

# RUN MODES
if __name__ == "__main__":
    # MODE 1: Single image (default)
    extract_and_combine_records(IMAGE_PATH, OUTPUT_FILE, Y_TOLERANCE)
    
    # MODE 2: Batch process 8 images (uncomment to use)
    """
    IMAGE_FOLDER = 'docs/images'  # Change to your folder with 8 images
    process_batch(IMAGE_FOLDER, 'output_combined')
    """