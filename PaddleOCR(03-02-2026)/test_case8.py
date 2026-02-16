# -*- coding: utf-8 -*-
"""
CRITICAL: Set Tesseract path BEFORE importing/using pytesseract functions
"""
import pytesseract
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

from PIL import Image
import pandas as pd
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font

def validate_and_normalize_path(image_path):
    """Validate and normalize the image path"""
    if not image_path:
        raise ValueError("Image path cannot be empty")

    if not os.path.isabs(image_path):
        image_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), image_path)

    if not os.path.exists(image_path):
        raise FileNotFoundError(f"Image not found at: {image_path}\n"
                              f"Current working directory: {os.getcwd()}\n"
                              f"Full resolved path: {os.path.abspath(image_path)}")
    return image_path

# --- IMAGE PATH ---
image_path = 'docs/images/Statista-FormF-RKJJ-NA-1929751_d.jpg'

try:
    image_path = validate_and_normalize_path(image_path)
except (FileNotFoundError, ValueError) as e:
    print(f"ERROR: {e}")
    sys.exit(1)

print("Loading and preprocessing image...")

# --- LOAD AND PREPROCESS IMAGE ---
img = Image.open(image_path)
img = img.convert('L')  # Grayscale
img = img.point(lambda x: 0 if x < 128 else 255, '1')  # Binarize
img = img.resize((img.width * 2, img.height * 2), Image.LANCZOS)  # Upscale 2x

# --- GET FULL TEXT LINES (for grouping) ---
text = pytesseract.image_to_string(img, lang='eng')
lines = [line.strip() for line in text.splitlines() if line.strip()]

if not lines:
    raise ValueError("No text detected in the image.")

print(f"✓ Total non-empty lines detected: {len(lines)}")

# --- GROUP LINES IN CHUNKS OF 4 ---
chunk_size = 4
groups = []
for i in range(0, len(lines), chunk_size):
    chunk = lines[i:i + chunk_size]
    words_in_group = []
    for line in chunk:
        words_in_group.extend(line.split())
    groups.append(words_in_group)

print(f"\n✓ Created {len(groups)} groups (each = up to {chunk_size} lines)")

# --- PAD TO MAX LENGTH ---
max_len = max(len(g) for g in groups) if groups else 0
padded_groups = [g + [''] * (max_len - len(g)) for g in groups]

# --- GET WORD-LEVEL CONFIDENCE ---
print("Running OCR with word-level confidence data...")
data = pytesseract.image_to_data(img, lang='eng', output_type=pytesseract.Output.DICT)

# Build word-confidence mapping
word_confidences = []
n_boxes = len(data['level'])
for i in range(n_boxes):
    conf = int(data['conf'][i])
    text_word = data['text'][i].strip()
    if conf >= 0 and text_word:
        word_confidences.append((text_word, conf))

# Flatten all words from groups
all_words_in_order = []
for group in padded_groups:
    for word in group:
        if word:
            all_words_in_order.append(word)

# Map words to confidence (position-based)
word_conf_map = {}
for i, word in enumerate(all_words_in_order):
    if i < len(word_confidences):
        word_conf_map[word] = word_confidences[i][1]
    else:
        word_conf_map[word] = 0

# --- CREATE EXCEL WITH HIGHLIGHTING ---
wb = Workbook()
ws = wb.active
red_font = Font(color="FF0000")

# Write rows with conditional formatting
for row_idx, row in enumerate(padded_groups, start=1):
    for col_idx, word in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=word)
        if word and word_conf_map.get(word, 100) < 99:
            ws.cell(row=row_idx, column=col_idx).font = red_font

# Save
output_excel = 'merged_4lines_per_row_highlighted.xlsx'
wb.save(output_excel)

print(f"\n✓ Saved {len(groups)} rows to: {output_excel}")
print(f"  Shape: {len(groups)} rows × {max_len} columns")
print("  Words with confidence < 99% are highlighted in RED.")