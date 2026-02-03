# test_case.py
import sys
import os
import logging

# --- Clean sys.path ---
for p in ["", ".", os.getcwd()]:
    if p in sys.path:
        sys.path.remove(p)

# --- Environment fixes ---
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
os.environ["DISABLE_MODEL_SOURCE_CHECK"] = "True"
logging.getLogger("ppocr").setLevel(logging.ERROR)

from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))  # Add current directory to path
from confidence_colors import get_confidence_color

def validate_and_normalize_path(image_path):
    """Validate and normalize the image path"""
    if not image_path:
        raise ValueError("Image path cannot be empty")

    # Convert to absolute path if relative
    if not os.path.isabs(image_path):
        image_path = os.path.join(os.path.dirname(__file__), image_path)

    # Check if file exists
    if not os.path.isfile(image_path):
        raise FileNotFoundError(f"Image not found: {image_path}")

    return image_path

# --- Image path ---
img_path = "docs/images/Statista-FormF-RKJJ-NA-1929751_d.jpg"  # Changed to relative path

try:
    img_path = validate_and_normalize_path(img_path)
except FileNotFoundError as e:
    print(f"Error: {e}")
    sys.exit(1)
except ValueError as e:
    print(f"Error: {e}")
    sys.exit(1)


# --- Initialize OCR ---
# ✅ Use only 'lang' to avoid parameter errors and enable high-accuracy model
ocr = PaddleOCR(lang="en",rec_char_type='en')
result = ocr.ocr(img_path)  # cls=True not needed; form is axis-aligned  # Correct API usage: using ocr() method

# --- Parse results and organize by rows ---
detections = []
if result is not None:
    # Handle PaddleOCR 3.x output format
    # Result is typically a list of prediction results
    for res in result:
        # Check if result has the expected structure
        if isinstance(res, dict) and 'boxes' in res and 'texts' in res and 'scores' in res:
            # Process each detected text element
            boxes = res['boxes']
            texts = res['texts']
            scores = res['scores']

            for i in range(len(texts)):
                box = boxes[i]
                text = texts[i]
                confidence = scores[i]

                if text and confidence is not None:
                    # Get the leftmost X-coordinate and top Y-coordinate of the bounding box
                    left_x = min(point[0] for point in box)
                    top_y = min(point[1] for point in box)  # Use Y-coordinate for row grouping
                    detections.append((left_x, top_y, text, confidence))
        elif isinstance(res, list) and len(res) > 0:
            # Fallback: older format or different structure
            items = res[0] if len(res) == 1 and isinstance(res[0], list) else res
            for item in items:
                if not item or len(item) < 2:
                    continue
                box = item[0]  # Bounding box: [[x1,y1], [x2,y2], [x3,y3], [x4,y4]]
                text_info = item[1]
                if isinstance(text_info, (list, tuple)) and len(text_info) >= 2:
                    try:
                        text = str(text_info[0]).strip()
                        confidence = float(text_info[1])
                        if text:
                            # Get the leftmost X-coordinate and top Y-coordinate of the bounding box
                            left_x = min(point[0] for point in box)
                            top_y = min(point[1] for point in box)  # Use Y-coordinate for row grouping
                            detections.append((left_x, top_y, text, confidence))
                    except (ValueError, TypeError, IndexError):
                        continue

# --- Sort by top_y first (top to bottom), then by left_x (left to right) ---
detections.sort(key=lambda x: (x[1], x[0]))

# --- Create Excel workbook ---
wb = Workbook()
ws = wb.active
ws.title = "OCR Results"

# Group detections by Y-coordinate to form rows as they appear in the image
from collections import defaultdict
rows_dict = defaultdict(list)
for x, y, text, confidence in detections:
    # Group words by similar Y-coordinates to form rows
    # Use Y-coordinate to group words into rows as they appear in the image
    rounded_y = round(y / 10) * 10  # Group by Y-coordinate with tolerance of 10 pixels
    rows_dict[rounded_y].append((x, text, confidence))

# Sort rows by Y-coordinate (top to bottom)
sorted_rows = sorted(rows_dict.items())

# Process each row and put words in separate columns
for row_key, row_words in sorted_rows:
    # Sort words in this row by X-coordinate (left to right)
    sorted_words = sorted(row_words, key=lambda item: item[0])

    # Create a row in Excel with each word in a separate column
    row_values = []
    row_confidences = []
    for x, text, confidence in sorted_words:
        row_values.append(text)
        row_confidences.append(confidence)

    # Add the row to the worksheet
    if row_values:  # Only add if there's data
        ws.append(row_values)

        # Get the current row index
        current_row_idx = ws.max_row

        # Apply background color to each cell based on confidence
        for col_idx, confidence in enumerate(row_confidences):
            # Get color based on confidence
            hex_color = get_confidence_color(confidence)

            # Apply background color to the cell
            cell = ws.cell(row=current_row_idx, column=col_idx + 1)  # +1 because Excel columns start at 1
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

            # Optionally, make low confidence text more visible
            if confidence < 0.70:
                # Make text bold for low confidence items
                cell.font = Font(bold=True)

# Save to Excel file
import datetime
timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
excel_file = f"output_highlighted_{timestamp}.xlsx"
wb.save(excel_file)

print(f"\nSUCCESS: Excel file saved with confidence-based highlights: {excel_file}")
print("Color scheme: Red (low confidence) to Yellow (medium) to Green (high confidence)")
print("Each row in Excel corresponds to a row in the image, with words in separate columns")