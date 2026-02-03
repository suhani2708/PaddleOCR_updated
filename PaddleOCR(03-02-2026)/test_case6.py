# test_case.py
import sys
import os
import logging

# --- Clean sys.path ---
for p in ["", ".", os.getcwd()]:
    if p in sys.path:
        sys.path.remove(p)

# --- Environment fixes ---
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
os.environ["DISABLE_MODEL_SOURCE_CHECK"] = "True"
logging.getLogger("ppocr").setLevel(logging.ERROR)

from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def validate_and_normalize_path(image_path):
    """Validate and normalize the image path"""
    if not image_path:
        raise ValueError("Image path cannot be empty")

    # Convert to absolute path if relative
    if not os.path.isabs(image_path):
        image_path = os.path.join(os.path.dirname(__file__), image_path)

    # Check if file exists
    if not os.path.isfile(image_path):
        raise FileNotFoundError(f"Image not found: {image_path}")

    return image_path

# --- Image path ---
img_path = "docs/images/Statista-FormF-RKJJ-NA-1929755_d.jpg"  # Changed to relative path

try:
    img_path = validate_and_normalize_path(img_path)
except FileNotFoundError as e:
    print(f"Error: {e}")
    sys.exit(1)
except ValueError as e:
    print(f"Error: {e}")
    sys.exit(1)

# --- Initialize OCR ---
ocr = PaddleOCR(lang="en", use_angle_cls=False, show_log=False, return_word_box=True)
result = ocr.ocr(img_path, cls=True)  # Correct API usage: using ocr() method

# --- Parse results with left-to-right sorting ---
detections = []
if result is not None:
    # Handle PaddleOCR output format: result[0] contains all detected text regions
    # Each region is: [bounding_box, (text, confidence)]
    # bounding_box is [[x1,y1], [x2,y2], [x3,y3], [x4,y4]]
    # text is a string, confidence is a float
    items = result[0] if len(result) == 1 and isinstance(result[0], list) else result
    for item in items:
        if not item or len(item) < 2:
            continue
        box = item[0]  # Bounding box: [[x1,y1], [x2,y2], [x3,y3], [x4,y4]]
        text_info = item[1]
        if isinstance(text_info, (list, tuple)) and len(text_info) >= 2:
            try:
                text = str(text_info[0]).strip()
                confidence = float(text_info[1])
                if text:
                    # Get the leftmost X-coordinate of the bounding box
                    left_x = min(point[0] for point in box)
                    detections.append((left_x, text, confidence))
            except (ValueError, TypeError, IndexError):
                continue

# --- Sort by left_x (left to right) ---
detections.sort(key=lambda x: x[0])

# --- Create Excel workbook ---
wb = Workbook()
ws = wb.active
ws.title = "OCR Results"

# Header
ws.append(["Text", "Confidence", "Low_Confidence"])

# Red fill for low-confidence rows
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

# Add data rows
row_index = 2  # Start from row 2 (row 1 is header)
for _, text, confidence in detections:
    low_flag = "YES" if confidence < 1.00 else "NO"
    ws.append([text, f"{confidence:.2f}", low_flag])
    
    # Highlight entire row if low confidence
    if low_flag == "YES":
        for col in range(1, 4):  # Columns A, B, C
            ws.cell(row=row_index, column=col).fill = red_fill
    row_index += 1

# Save to Excel file
excel_file = "output_highlighted6.xlsx"
wb.save(excel_file)

print(f"\n✅ Excel file saved with highlights: {excel_file}")