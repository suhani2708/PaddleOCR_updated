"""
Excel exporter for PaddleOCR results.
Converts OCR results to Excel format with confidence-based highlighting.
"""

from typing import Dict, List, Tuple, Any, Optional
from pathlib import Path
import logging
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from paddleocr import PaddleOCR
import cv2
import numpy as np

from text_row_detector import detect_text_rows_from_bounding_boxes
from confidence_colors import get_confidence_color
from ocr_utils import extract_word_data


class PaddleOCRExcelExporter:
    """
    A class to export PaddleOCR results to Excel format with confidence-based highlighting.
    
    This class provides a complete workflow for processing images with OCR, organizing
    the results into rows, and exporting them to Excel with visual indicators for
    low-confidence text that may require human verification.
    
    Attributes:
        ocr (PaddleOCR): The PaddleOCR instance for performing OCR
        confidence_threshold (float): Threshold below which text is considered low confidence
        enable_orientation (bool): Whether to enable text orientation detection
    """
    
    def __init__(self, lang: str = 'en', confidence_threshold: float = 1.00,
                 enable_orientation: bool = True, return_word_box: bool = True,
                 y_tolerance: int = 10, reading_direction: str = 'ltr'):
        """
        Initialize the PaddleOCRExcelExporter.

        Args:
            lang: Language for OCR processing (default: 'en')
            confidence_threshold: Threshold below which text is highlighted (default: 1.00)
            enable_orientation: Whether to enable text orientation detection (default: True)
            return_word_box: Whether to return word-level bounding boxes (default: True)
            y_tolerance: Tolerance for row detection based on Y-coordinates (default: 10)
            reading_direction: Reading direction for text ('ltr', 'rtl', 'auto') (default: 'ltr')
        """
        # Input validation
        if not isinstance(lang, str):
            raise TypeError("Language must be a string")

        if not (0.0 <= confidence_threshold <= 1.0):
            raise ValueError(f"Confidence threshold must be between 0.0 and 1.0, got {confidence_threshold}")

        if not isinstance(enable_orientation, bool):
            raise TypeError("enable_orientation must be a boolean")

        if not isinstance(return_word_box, bool):
            raise TypeError("return_word_box must be a boolean")

        if not isinstance(y_tolerance, int) or y_tolerance < 0:
            raise ValueError(f"Y tolerance must be a non-negative integer, got {y_tolerance}")

        if reading_direction not in ['ltr', 'rtl', 'auto']:
            raise ValueError(f"Reading direction must be 'ltr', 'rtl', or 'auto', got {reading_direction}")

        # Store configuration
        self.lang = lang
        self.confidence_threshold = confidence_threshold
        self.enable_orientation = enable_orientation
        self.return_word_box = return_word_box
        self.y_tolerance = y_tolerance
        self.reading_direction = reading_direction

        try:
            # Initialize PaddleOCR with specified parameters
            # use_angle_cls enables text line orientation classification (rotated text detection)
            # use_textline_orientation enables text line orientation detection (if supported in newer versions)
            self.ocr = PaddleOCR(
                lang=lang,
                use_angle_cls=enable_orientation,
                show_log=False,
                return_word_box=return_word_box
            )
        except Exception as e:
            self.logger.error(f"Failed to initialize PaddleOCR: {str(e)}")
            raise

        # Set up logging
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)

    def set_confidence_threshold(self, threshold: float) -> None:
        """
        Set the confidence threshold for highlighting low-confidence text.

        Args:
            threshold: New confidence threshold value (between 0.0 and 1.0)
        """
        if not (0.0 <= threshold <= 1.0):
            raise ValueError(f"Confidence threshold must be between 0.0 and 1.0, got {threshold}")

        self.confidence_threshold = threshold
        self.logger.info(f"Confidence threshold updated to {threshold}")

    def set_y_tolerance(self, y_tolerance: int) -> None:
        """
        Set the Y-coordinate tolerance for row detection.

        Args:
            y_tolerance: New Y-coordinate tolerance value (non-negative integer)
        """
        if not isinstance(y_tolerance, int) or y_tolerance < 0:
            raise ValueError(f"Y tolerance must be a non-negative integer, got {y_tolerance}")

        self.y_tolerance = y_tolerance
        self.logger.info(f"Y tolerance updated to {y_tolerance}")

    def set_reading_direction(self, reading_direction: str) -> None:
        """
        Set the reading direction for text processing.

        Args:
            reading_direction: New reading direction ('ltr', 'rtl', or 'auto')
        """
        if reading_direction not in ['ltr', 'rtl', 'auto']:
            raise ValueError(f"Reading direction must be 'ltr', 'rtl', or 'auto', got {reading_direction}")

        self.reading_direction = reading_direction
        self.logger.info(f"Reading direction updated to {reading_direction}")

    def get_config(self) -> Dict[str, Any]:
        """
        Get the current configuration settings.

        Returns:
            Dictionary containing all configuration parameters
        """
        return {
            'lang': self.lang,
            'confidence_threshold': self.confidence_threshold,
            'enable_orientation': self.enable_orientation,
            'return_word_box': self.return_word_box,
            'y_tolerance': self.y_tolerance,
            'reading_direction': self.reading_direction
        }

    def preprocess_image_for_orientation(self, image_path: str) -> str:
        """
        Preprocess an image to handle rotation and improve OCR results.

        Args:
            image_path: Path to the input image

        Returns:
            Path to the preprocessed image (may be the same as input if no preprocessing needed)
        """
        try:
            # Load the image
            img = cv2.imread(image_path)
            if img is None:
                raise ValueError(f"Could not load image: {image_path}")

            # If orientation detection is enabled, we rely on PaddleOCR's built-in capabilities
            # But we can still do basic preprocessing to improve results
            # Convert to grayscale for potential rotation detection
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            # Apply basic preprocessing to enhance text
            # Apply Gaussian blur to reduce noise
            blurred = cv2.GaussianBlur(gray, (3, 3), 0)

            # Apply threshold to get a binary image (useful for text detection)
            _, thresh = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

            # Return original image path since PaddleOCR handles orientation internally
            # The preprocessing is mainly handled by PaddleOCR when use_angle_cls=True
            return image_path

        except Exception as e:
            self.logger.warning(f"Image preprocessing failed: {str(e)}, using original image")
            return image_path
    
    def process_image(self, image_path: str) -> List[Dict[str, Any]]:
        """
        Process an image with OCR and extract structured word data.

        Args:
            image_path: Path to the image file to process

        Returns:
            List of dictionaries containing word data with text, coordinates, and confidence
        """
        self.logger.info(f"Processing image: {image_path}")

        # Validate image path
        if not isinstance(image_path, str):
            raise TypeError("Image path must be a string")

        if not os.path.exists(image_path):
            raise FileNotFoundError(f"Image not found: {image_path}")

        # Check if image is readable
        if not os.access(image_path, os.R_OK):
            raise PermissionError(f"Cannot read image file: {image_path}")

        # Validate image format
        try:
            image_extension = Path(image_path).suffix.lower()
            valid_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.gif']
            if image_extension not in valid_extensions:
                raise ValueError(f"Unsupported image format: {image_extension}. Supported formats: {valid_extensions}")

            # Try to load the image with OpenCV to verify it's a valid image
            img = cv2.imread(image_path)
            if img is None:
                raise ValueError(f"Cannot read image file: {image_path}. File may be corrupted or in unsupported format.")
        except Exception as e:
            self.logger.error(f"Image format validation failed: {str(e)}")
            raise

        try:
            # Preprocess image for better orientation handling
            processed_image_path = self.preprocess_image_for_orientation(image_path)

            # Run OCR on image with text orientation detection enabled
            result = self.ocr.ocr(
                processed_image_path,
                cls=self.enable_orientation
            )

            # Handle OCR failures (empty results)
            if not result:
                self.logger.warning(f"No text detected in image: {image_path}")
                return []

            # Extract word-level data
            word_data = extract_word_data(result)

            if not word_data:
                self.logger.warning(f"No word data extracted from image: {image_path}")
                return []

            self.logger.info(f"Extracted {len(word_data)} words from image")

            # Organize into rows based on Y-coordinates
            detections = [(item['x'], item['text'], item['confidence'], item['y']) for item in word_data]
            rows = detect_text_rows_from_bounding_boxes(detections)

            # Convert to list of rows, each containing a list of words
            rows_list = []
            for y_coord, words in sorted(rows.items()):
                row_data = []
                for x_pos, text, confidence in words:
                    # Validate confidence values
                    if not (0.0 <= confidence <= 1.0):
                        self.logger.warning(f"Invalid confidence value: {confidence}. Setting to 0.0")
                        confidence = 0.0

                    row_data.append({
                        'text': text,
                        'confidence': confidence,
                        'x': x_pos,
                        'y': y_coord
                    })
                rows_list.append(row_data)

            self.logger.info(f"Organized words into {len(rows_list)} rows")

            return rows_list

        except Exception as e:
            self.logger.error(f"Error processing image {image_path}: {str(e)}")
            raise
    
    def create_excel_workbook(self, rows_data: List[List[Dict[str, Any]]]) -> Workbook:
        """
        Create an Excel workbook from processed OCR data.

        Args:
            rows_data: List of rows, each containing a list of word dictionaries
                      Each word dictionary has 'text', 'confidence', 'x', 'y' keys

        Returns:
            openpyxl.Workbook object with formatted data
        """
        self.logger.info("Creating Excel workbook")

        # Validate input data
        if not isinstance(rows_data, list):
            raise TypeError("rows_data must be a list")

        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "OCR Results"

            # Add header row explaining the color scheme
            header = ["Word 1", "Word 2", "Word 3", "Word 4", "..."]  # Dynamic based on content
            ws.append(header)

            # Add color scheme explanation in a separate row
            color_explanation = ["Color Guide:", "Green = High Confidence", "Yellow = Medium Confidence", "Red = Low Confidence", ""]
            ws.append(color_explanation)

            # Add an empty row before the actual data
            ws.append([])

            # Process each row of detected text
            for row_data in rows_data:
                if not isinstance(row_data, list):
                    self.logger.warning(f"Skipping invalid row data: {row_data}")
                    continue

                # Create a new Excel row with each word in a separate column
                row_values = []
                for word_info in row_data:
                    if not isinstance(word_info, dict) or 'text' not in word_info:
                        self.logger.warning(f"Skipping invalid word data: {word_info}")
                        continue

                    text = word_info['text']
                    row_values.append(text)

                # Add the row to the worksheet
                if row_values:  # Only add if there's data
                    ws.append(row_values)

                    # Get the current row index
                    current_row_idx = ws.max_row

                    # Apply background color based on confidence using get_confidence_color function
                    for col_idx, word_info in enumerate(row_data):
                        if 'confidence' not in word_info:
                            continue  # Skip if no confidence value

                        confidence = word_info['confidence']

                        # Validate confidence value
                        if not isinstance(confidence, (int, float)) or not (0.0 <= confidence <= 1.0):
                            self.logger.warning(f"Invalid confidence value: {confidence}. Using 0.0")
                            confidence = 0.0

                        try:
                            # Get the appropriate color based on confidence
                            hex_color = get_confidence_color(confidence)

                            # Create a fill with the appropriate gradient color
                            cell_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

                            # Apply the fill to the cell
                            cell = ws.cell(row=current_row_idx, column=col_idx + 1)  # +1 because Excel columns start at 1
                            cell.fill = cell_fill
                        except Exception as color_error:
                            self.logger.warning(f"Could not apply color to cell: {color_error}")
                            # Continue without coloring if color application fails

            # Auto-adjust column widths for readability
            try:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    # Adjust width based on content, with reasonable bounds
                    adjusted_width = min(max(max_length + 2, 10), 50)  # Min 10, Max 50, based on content
                    ws.column_dimensions[column_letter].width = adjusted_width
            except Exception as width_error:
                self.logger.warning(f"Could not adjust column widths: {width_error}")

        except Exception as e:
            self.logger.error(f"Error creating Excel workbook: {str(e)}")
            raise

        self.logger.info("Excel workbook created successfully with confidence-based coloring")
        return wb
    
    def export_to_excel(self, image_path: str, excel_path: str) -> str:
        """
        Complete workflow: Process image with OCR and export results to Excel.

        Args:
            image_path: Path to the input image file
            excel_path: Path for the output Excel file

        Returns:
            Path to the created Excel file
        """
        self.logger.info(f"Starting export process: {image_path} -> {excel_path}")

        # Validate input parameters
        if not isinstance(image_path, str):
            raise TypeError("Image path must be a string")

        if not isinstance(excel_path, str):
            raise TypeError("Excel path must be a string")

        # Validate output path
        excel_path_obj = Path(excel_path)
        output_dir = excel_path_obj.parent

        # Check if output directory is writable
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            # Create a temporary file to test write permissions
            test_file = output_dir / ".write_test"
            test_file.touch()
            test_file.unlink()  # Remove the test file
        except Exception as e:
            raise PermissionError(f"Cannot write to output directory: {output_dir}. Error: {str(e)}")

        # Check if output file extension is valid
        if excel_path_obj.suffix.lower() not in ['.xlsx', '.xls']:
            raise ValueError(f"Invalid Excel file extension: {excel_path_obj.suffix}. Use .xlsx or .xls")

        try:
            # Process the image
            rows_data = self.process_image(image_path)

            # Create Excel workbook
            wb = self.create_excel_workbook(rows_data)

            # Save workbook with error handling
            try:
                wb.save(excel_path)
            except PermissionError:
                raise PermissionError(f"Permission denied when saving Excel file: {excel_path}")
            except OSError as e:
                if "Disk full" in str(e) or e.errno == 28:  # errno 28 = No space left on device
                    raise OSError(f"Not enough disk space to save Excel file: {excel_path}")
                else:
                    raise OSError(f"Error saving Excel file: {str(e)}")
            except Exception as e:
                raise Exception(f"Unexpected error saving Excel file: {str(e)}")

            self.logger.info(f"Export completed successfully: {excel_path}")
            return excel_path

        except FileNotFoundError:
            self.logger.error(f"Image file not found: {image_path}")
            raise
        except PermissionError:
            self.logger.error(f"Permission error during export: {image_path} -> {excel_path}")
            raise
        except OSError as e:
            self.logger.error(f"OS error during export: {str(e)}")
            raise
        except Exception as e:
            self.logger.error(f"Error during export process: {str(e)}")
            raise
    
    def process_image_for_columns(self, image_path: str) -> List[List[str]]:
        """
        Process an image and return data organized by rows and columns for Excel.

        Args:
            image_path: Path to the image file to process

        Returns:
            List of rows, each containing a list of text strings
        """
        self.logger.info(f"Processing image for column-based export: {image_path}")

        # Validate image path
        if not isinstance(image_path, str):
            raise TypeError("Image path must be a string")

        if not os.path.exists(image_path):
            raise FileNotFoundError(f"Image not found: {image_path}")

        # Check if image is readable
        if not os.access(image_path, os.R_OK):
            raise PermissionError(f"Cannot read image file: {image_path}")

        # Validate image format
        try:
            image_extension = Path(image_path).suffix.lower()
            valid_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.gif']
            if image_extension not in valid_extensions:
                raise ValueError(f"Unsupported image format: {image_extension}. Supported formats: {valid_extensions}")

            # Try to load the image with OpenCV to verify it's a valid image
            img = cv2.imread(image_path)
            if img is None:
                raise ValueError(f"Cannot read image file: {image_path}. File may be corrupted or in unsupported format.")
        except Exception as e:
            self.logger.error(f"Image format validation failed: {str(e)}")
            raise

        try:
            # Preprocess image for better orientation handling
            processed_image_path = self.preprocess_image_for_orientation(image_path)

            # Run OCR on image with text orientation detection enabled
            result = self.ocr.ocr(
                processed_image_path,
                cls=self.enable_orientation
            )

            # Handle OCR failures (empty results)
            if not result:
                self.logger.warning(f"No text detected in image: {image_path}")
                return []

            # Extract word-level data
            word_data = extract_word_data(result)

            if not word_data:
                self.logger.warning(f"No word data extracted from image: {image_path}")
                return []

            self.logger.info(f"Extracted {len(word_data)} words from image")

            # Organize into rows based on Y-coordinates
            detections = [(item['x'], item['text'], item['confidence'], item['y']) for item in word_data]
            rows = detect_text_rows_from_bounding_boxes(detections)

            # Convert to list of rows, each containing a list of text strings
            rows_list = []
            for y_coord, words in sorted(rows.items()):
                row_texts = [text for x_pos, text, confidence in words]
                rows_list.append(row_texts)

            self.logger.info(f"Organized words into {len(rows_list)} rows")

            return rows_list

        except Exception as e:
            self.logger.error(f"Error processing image {image_path}: {str(e)}")
            raise
    
    def create_excel_workbook_for_columns(self, rows_data: List[List[str]]) -> Workbook:
        """
        Create an Excel workbook with each word in a separate column per row.

        Args:
            rows_data: List of rows, each containing a list of text strings

        Returns:
            openpyxl.Workbook object with formatted data
        """
        self.logger.info("Creating Excel workbook with words in separate columns")

        # Validate input data
        if not isinstance(rows_data, list):
            raise TypeError("rows_data must be a list")

        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "OCR Results"

            # Add data rows (each word in a separate column)
            for row_data in rows_data:
                if not isinstance(row_data, list):
                    self.logger.warning(f"Skipping invalid row data: {row_data}")
                    continue

                # Convert all items to strings to ensure compatibility
                row_values = []
                for item in row_data:
                    try:
                        row_values.append(str(item))
                    except Exception as e:
                        self.logger.warning(f"Could not convert item to string: {item}, error: {e}")
                        row_values.append("")  # Use empty string as fallback

                if row_values:  # Only append if there are values
                    ws.append(row_values)

            # Auto-adjust column widths
            try:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max(max_length + 2, 10), 50)  # Min 10, cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
            except Exception as width_error:
                self.logger.warning(f"Could not adjust column widths: {width_error}")

        except Exception as e:
            self.logger.error(f"Error creating Excel workbook for columns: {str(e)}")
            raise

        self.logger.info("Excel workbook created successfully with words in separate columns")
        return wb
    
    def export_to_excel_with_confidence_colors(self, image_path: str, excel_path: str) -> str:
        """
        Export OCR results to Excel with confidence-based cell coloring.

        Args:
            image_path: Path to the input image file
            excel_path: Path for the output Excel file

        Returns:
            Path to the created Excel file
        """
        self.logger.info(f"Starting export process with confidence colors: {image_path} -> {excel_path}")

        # Validate input parameters
        if not isinstance(image_path, str):
            raise TypeError("Image path must be a string")

        if not isinstance(excel_path, str):
            raise TypeError("Excel path must be a string")

        # Validate output path
        excel_path_obj = Path(excel_path)
        output_dir = excel_path_obj.parent

        # Check if output directory is writable
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            # Create a temporary file to test write permissions
            test_file = output_dir / ".write_test"
            test_file.touch()
            test_file.unlink()  # Remove the test file
        except Exception as e:
            raise PermissionError(f"Cannot write to output directory: {output_dir}. Error: {str(e)}")

        # Check if output file extension is valid
        if excel_path_obj.suffix.lower() not in ['.xlsx', '.xls']:
            raise ValueError(f"Invalid Excel file extension: {excel_path_obj.suffix}. Use .xlsx or .xls")

        try:
            # Process the image to get word data
            result = self.ocr.ocr(
                image_path,
                cls=self.enable_orientation
            )

            # Handle OCR failures (empty results)
            if not result:
                self.logger.warning(f"No text detected in image: {image_path}")
                # Create an empty workbook anyway
                wb = Workbook()
                ws = wb.active
                ws.title = "OCR Results"
                ws.append(["No text detected"])
            else:
                word_data = extract_word_data(result)

                if not word_data:
                    self.logger.warning(f"No word data extracted from image: {image_path}")
                    # Create an empty workbook anyway
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "OCR Results"
                    ws.append(["No word data extracted"])
                else:
                    # Organize into rows based on Y-coordinates
                    detections = [(item['x'], item['text'], item['confidence'], item['y']) for item in word_data]
                    rows = detect_text_rows_from_bounding_boxes(detections)

                    # Create workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "OCR Results"

                    # Define fills for confidence-based coloring
                    for y_coord, words in sorted(rows.items()):
                        row_values = []
                        row_colors = []

                        for x_pos, text, confidence in words:
                            row_values.append(text)

                            # Validate confidence value
                            if not isinstance(confidence, (int, float)) or not (0.0 <= confidence <= 1.0):
                                self.logger.warning(f"Invalid confidence value: {confidence}. Using 0.0")
                                confidence = 0.0

                            # Get color based on confidence
                            try:
                                hex_color = get_confidence_color(confidence)
                                row_colors.append(hex_color)
                            except Exception as color_error:
                                self.logger.warning(f"Could not get color for confidence {confidence}: {color_error}")
                                row_colors.append("FFFFFF")  # Default to white

                        # Append the row to the worksheet
                        ws.append(row_values)

                        # Apply colors to the cells in this row
                        current_row_idx = ws.max_row
                        for col_idx, color in enumerate(row_colors, start=1):
                            try:
                                cell = ws.cell(row=current_row_idx, column=col_idx)
                                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                                cell.fill = fill
                            except Exception as cell_error:
                                self.logger.warning(f"Could not apply color to cell at row {current_row_idx}, col {col_idx}: {cell_error}")

            # Auto-adjust column widths
            try:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max(max_length + 2, 10), 50)  # Min 10, cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
            except Exception as width_error:
                self.logger.warning(f"Could not adjust column widths: {width_error}")

            # Save workbook with error handling
            try:
                wb.save(excel_path)
            except PermissionError:
                raise PermissionError(f"Permission denied when saving Excel file: {excel_path}")
            except OSError as e:
                if "Disk full" in str(e) or e.errno == 28:  # errno 28 = No space left on device
                    raise OSError(f"Not enough disk space to save Excel file: {excel_path}")
                else:
                    raise OSError(f"Error saving Excel file: {str(e)}")
            except Exception as e:
                raise Exception(f"Unexpected error saving Excel file: {str(e)}")

            self.logger.info(f"Export with confidence colors completed successfully: {excel_path}")
            return excel_path

        except FileNotFoundError:
            self.logger.error(f"Image file not found: {image_path}")
            raise
        except PermissionError:
            self.logger.error(f"Permission error during export: {image_path} -> {excel_path}")
            raise
        except OSError as e:
            self.logger.error(f"OS error during export: {str(e)}")
            raise
        except Exception as e:
            self.logger.error(f"Error during export process with confidence colors: {str(e)}")
            raise