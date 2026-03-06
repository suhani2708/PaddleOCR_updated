import os
import cv2
import numpy as np
import openpyxl
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from paddleocr import PaddleOCR
# ─────────────────────────────────────────────────────────────────────────────
# CONFIG  ←  Only edit these two lines
# ─────────────────────────────────────────────────────────────────────────────
IMAGE_PATH = r"C:\Users\91909\Desktop\github new repository\new repo\ftb(30-12-2025)\PaddleOCR_updated\PaddleOCR(03-02-2026)\docs\images\Statista-FormF-RKJJ-NA-1929752_d_new.jpg"
# Output Excel is named after the image file — saved in the same folder as this script
OUTPUT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           os.path.splitext(os.path.basename(IMAGE_PATH))[0] + ".xlsx")
# ─────────────────────────────────────────────────────────────────────────────
# THRESHOLDS
# ─────────────────────────────────────────────────────────────────────────────
LOW_CONF_THRESHOLD    = 0.65
MID_CONF_THRESHOLD    = 0.88
IOU_MATCH_THRESHOLD   = 0.15
VERIFY_CONF_THRESHOLD = 0.88
CROP_PAD              = 10
MAX_WORKERS           = 8
MAX_SIDE              = 3500
MAX_SCALE             = 2.0
# Punctuation characters we track for the "missing punctuation" RED highlight
PUNCT_CHARS = set('.,;:!?\'"-–—()[]{}/\\@#%&*+=<>~`')
# ─────────────────────────────────────────────────────────────────────────────
# SINGLETON OCR ENGINE
# ─────────────────────────────────────────────────────────────────────────────
_ocr_engine = None
def get_ocr_engine():
    global _ocr_engine
    if _ocr_engine is None:
        _ocr_engine = PaddleOCR(
            use_angle_cls=True,
            lang='en',
            use_gpu=False,
            show_log=False,
            det_db_thresh=0.05,
            det_db_box_thresh=0.1,
            det_db_unclip_ratio=1.6,
            det_limit_side_len=4800,
        )
    return _ocr_engine
# ─────────────────────────────────────────────────────────────────────────────
# IMAGE PREPROCESSING
# ─────────────────────────────────────────────────────────────────────────────
def preprocess_full(path):
    img = cv2.imread(path)
    if img is None:
        raise FileNotFoundError(
            f"\n  Cannot open image: {path}\n"
            f"  Please check the IMAGE_PATH setting at the top of this script."
        )
    h, w = img.shape[:2]
    scale = min(MAX_SCALE, MAX_SIDE / max(h, w))
    img = cv2.resize(img, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    kernel = np.array([[0, -1, 0],
                       [-1,  5, -1],
                       [0, -1, 0]])
    return cv2.filter2D(gray, -1, kernel)
# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def iou(boxA, boxB):
    xA = max(boxA[0], boxB[0]); yA = max(boxA[1], boxB[1])
    xB = min(boxA[2], boxB[2]); yB = min(boxA[3], boxB[3])
    inter = max(0, xB - xA) * max(0, yB - yA)
    if inter == 0:
        return 0.0
    aA = (boxA[2] - boxA[0]) * (boxA[3] - boxA[1])
    aB = (boxB[2] - boxB[0]) * (boxB[3] - boxB[1])
    return inter / float(aA + aB - inter)
def alphanum_len(text):
    return len(re.sub(r'[^a-zA-Z0-9]', '', text))
def punct_count(text):
    return sum(1 for c in text if c in PUNCT_CHARS)
def fix_punctuation(text):
    if not text:
        return text
    text = re.sub(r'(\d),(\d)', r'\1.\2', text)
    text = text.replace(' ,', ' .')
    text = text.replace(', ', ' . ')
    text = re.sub(r'(\d{1,2}),(\d{1,2}),(\d{4})', r'\1.\2.\3', text)
    if text.endswith(','):
        text = text[:-1] + '.'
    return text.strip()
def make_variants(crop_gray):
    h = crop_gray.shape[0]
    scale = max(1.0, 80.0 / h)
    if scale > 1.0:
        crop_gray = cv2.resize(crop_gray, None, fx=scale, fy=scale,
                               interpolation=cv2.INTER_CUBIC)
    if crop_gray.shape[0] > 30:
        crop_gray = cv2.GaussianBlur(crop_gray, (3, 3), 0.5)
    v1 = crop_gray
    _, v2 = cv2.threshold(crop_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return [v1, v2]
# ─────────────────────────────────────────────────────────────────────────────
# CROP HELPER
# ─────────────────────────────────────────────────────────────────────────────
def get_crop(full_img, item):
    x1, y1, x2, y2 = [int(v) for v in item["bbox"]]
    H, W = full_img.shape[:2]
    cx1 = max(0, x1 - CROP_PAD)
    cy1 = max(0, y1 - CROP_PAD)
    cx2 = min(W, x2 + CROP_PAD)
    cy2 = min(H, y2 + CROP_PAD)
    return full_img[cy1:cy2, cx1:cx2]
# ─────────────────────────────────────────────────────────────────────────────
# PATTERN-BASED OCR ERROR DETECTION
# Detects missing characters using known data format patterns.
# These are the exact error types seen in OCR output vs image ground truth.
# ─────────────────────────────────────────────────────────────────────────────
def detect_pattern_errors(text):
    """
    Returns (has_error, reason_string).
    Checks for these known OCR failure patterns:
    1. Completely unread  →  "???"
    2. Broken phone paren →  "(NNN<digits>" missing closing ')'
    3. Missing 'x' in phone extension  →  NNN.NNN.NNNNNNNN (7+ digits after last dot/dash)
    4. Missing 'x' in dash-style extension  →  NNN-NNN-NNNNNNN (no x)
    5. Missing space around '&'  →  word&word or word& word
    6. Merged CamelCase words (missing comma/space)  →  TwoWords merged without separator
    """
    if not text or not isinstance(text, str):
        return False, None
    t = text.strip()
    # ── 1. Completely unread ──────────────────────────────────────────────────
    if t == '???':
        return True, "Region completely unread by OCR"
    # ── 2. Broken opening parenthesis in phone — '(' never closed ─────────────
    #    e.g. "(774820-7914"  should be  "(774)820-7914"
    if re.match(r'^\(\d{3}\d', t) and ')' not in t:
        return True, f"Missing ')' in phone number — extracted: '{t}'"
    # ── 3. Missing 'x' in dot/dash phone extension ───────────────────────────
    #    e.g. "654.655.93737946"  →  "654.655.9373x7946"
    #    e.g. "970.966.3554165"   →  "970.966.3554x165"
    #    e.g. "297.040.23642365"  →  "297.040.2364x2365"
    #    Pattern: three digit groups separated by dots/dashes,
    #             last group has 7+ digits (should have been split by 'x')
    if re.search(r'\d{3}[.\-]\d{3}[.\-]\d{7,}', t):
        return True, f"Missing 'x' in phone extension — extracted: '{t}'"
    # ── 4. Missing 'x' in dash-only style  ───────────────────────────────────
    #    e.g. "001-692-395-8914975"  →  "001-692-395-8914x975"
    #    Only flag when there is no 'x' already present
    if re.search(r'\d{3}-\d{3}-\d{4}\d{3,}', t) and 'x' not in t:
        return True, f"Missing 'x' in phone extension — extracted: '{t}'"
    # ── 5. Missing space around '&' ───────────────────────────────────────────
    #    e.g. "Hospital &Health Care"  →  "Hospital & Health Care"
    #    Exclude URLs, emails, SNS codes (contain / @ # %)
    if re.search(r'&\S', t) or re.search(r'\S&\S', t):
        if not re.search(r'[/@#%+]', t) and len(t) < 60:
            return True, f"Missing space around '&' — extracted: '{t}'"
    # ── 6. Merged CamelCase words — missing comma/space separator ─────────────
    #    e.g. "BarnesBarton and Mclaughlin"  →  "Barnes,Barton and Mclaughlin"
    #    e.g. "Maxwell.Guerrero"  is fine (dot is valid separator)
    #    Heuristic: lowercase letter immediately followed by uppercase letter,
    #    in a short string that looks like a name (no digits, slashes, @)
    if re.search(r'[a-z][A-Z]', t):
        if not re.search(r'[/\d@.\-_]', t) and len(t) < 50:
            return True, f"Missing separator between words (merged CamelCase) — extracted: '{t}'"
    return False, None
# ─────────────────────────────────────────────────────────────────────────────
# PASS 3 — verify word (recover missing letters in low-confidence words)
# ─────────────────────────────────────────────────────────────────────────────
def verify_word(full_img, item):
    conf = item.get("conf", 0)
    text = item["text"].strip()
    # High-confidence words: just fix punctuation, skip re-OCR
    if conf >= VERIFY_CONF_THRESHOLD:
        item["text"] = fix_punctuation(text)
        return False, item["text"]
    crop = get_crop(full_img, item)
    if crop.size == 0:
        return False, text
    variants = make_variants(crop)
    orig_len = alphanum_len(text)
    best_text = text
    best_len = orig_len
    ocr = get_ocr_engine()
    for variant in variants:
        try:
            # ── ONLY CHANGE: det=False skips slow detection — crop bbox already known ──
            res = ocr.ocr(variant, det=False, cls=True)
            if not res or res[0] is None:
                continue
            parts = [det[0].strip() for det in res[0] if det and det[0]]
            joined = " ".join(parts).strip()
            if not joined:
                continue
            vlen = alphanum_len(joined)
            if vlen > best_len:
                best_len = vlen
                best_text = joined
        except Exception:
            continue
    changed = abs(best_len - orig_len) > 1
    best_text = fix_punctuation(best_text)
    item["text"] = best_text
    if changed:
        if best_len > orig_len:
            item["missing_letters"] = True
        else:
            item["extra_garbage"] = True
    return changed, best_text
# ─────────────────────────────────────────────────────────────────────────────
# PASS 4 — detect missing punctuation (image has punct, Excel cell does not)
# ─────────────────────────────────────────────────────────────────────────────
def check_missing_punct(full_img, item):
    text = item["text"]
    orig_pc = punct_count(text)
    crop = get_crop(full_img, item)
    if crop.size == 0:
        item["missing_punct"] = False
        return
    variants = make_variants(crop)
    best_text = text
    best_pc = orig_pc
    ocr = get_ocr_engine()
    for variant in variants:
        try:
            # ── ONLY CHANGE: det=False skips slow detection — crop bbox already known ──
            res = ocr.ocr(variant, det=False, cls=True)
            if not res or res[0] is None:
                continue
            parts = [det[0].strip() for det in res[0] if det and det[0]]
            joined = " ".join(parts).strip()
            if not joined:
                continue
            pc = punct_count(joined)
            if pc > best_pc:
                best_pc = pc
                best_text = joined
        except Exception:
            continue
    if best_pc > orig_pc:
        item["missing_punct"] = True
        item["punct_reocr"] = best_text
    else:
        item["missing_punct"] = False
        item["punct_reocr"] = None
# ─────────────────────────────────────────────────────────────────────────────
# MAIN OCR PIPELINE
# ─────────────────────────────────────────────────────────────────────────────
def run_ocr(img):
    ocr = get_ocr_engine()
    items = []
    ocr_boxes = []
    # ── Pass 1: Full-image OCR ────────────────────────────────────────────────
    print("Pass 1: Full OCR...")
    result = ocr.ocr(img, cls=True)
    if result and result[0]:
        for line in result:
            for det in line:
                bbox, (text, conf) = det
                ys = [p[1] for p in bbox]
                xs = [p[0] for p in bbox]
                x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)
                clean = re.sub(r'(\w+)\.(\s+[A-Z])', r'\1,\2', text.strip())
                items.append({
                    "text":            clean,
                    "conf":            conf,
                    "x_start":         x1,
                    "y_center":        np.mean(ys),
                    "height":          y2 - y1,
                    "bbox":            (x1, y1, x2, y2),
                    "missed":          False,
                    "missing_letters": False,
                    "extra_garbage":   False,
                    "missing_punct":   False,
                    "punct_reocr":     None,
                    "pattern_error":   False,
                    "pattern_reason":  None,
                })
                ocr_boxes.append((x1, y1, x2, y2))
    # ── Pass 2: Detect regions OCR missed entirely ────────────────────────────
    print("Pass 2: Detecting missed regions...")
    det_result = ocr.ocr(img, rec=False, cls=False)
    if det_result and det_result[0]:
        for det_line in det_result:
            for bbox in det_line:
                xs = [p[0] for p in bbox]
                ys = [p[1] for p in bbox]
                x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)
                det_box = (x1, y1, x2, y2)
                if not any(iou(det_box, ob) > IOU_MATCH_THRESHOLD for ob in ocr_boxes):
                    items.append({
                        "text":            "???",
                        "conf":            -1,
                        "x_start":         x1,
                        "y_center":        np.mean(ys),
                        "height":          y2 - y1,
                        "bbox":            det_box,
                        "missed":          True,
                        "missing_letters": False,
                        "extra_garbage":   False,
                        "missing_punct":   False,
                        "punct_reocr":     None,
                        "pattern_error":   False,
                        "pattern_reason":  None,
                    })
    # ── Pass 3: Parallel re-verification for low-confidence words ─────────────
    to_verify = [
        (i, item) for i, item in enumerate(items)
        if not item["missed"] and item.get("conf", 1.0) < VERIFY_CONF_THRESHOLD
    ]
    skipped = len(items) - len(to_verify) - sum(1 for it in items if it["missed"])
    print(f"Pass 3: Verifying {len(to_verify)} low-confidence words "
          f"(skipping {skipped} high-confidence)...")
    def _verify_task(args):
        return verify_word(img, args[1])
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(_verify_task, arg): arg for arg in to_verify}
        done = 0
        for future in as_completed(futures):
            done += 1
            if done % 50 == 0:
                print(f"  ... {done}/{len(to_verify)}")
            try:
                future.result()
            except Exception as e:
                print(f"  [warn] verify error: {e}")
    # ── Pass 4: Parallel punctuation-missing detection ────────────────────────
    punct_candidates = [
        (i, item) for i, item in enumerate(items)
        if not item["missed"]
        and (item.get("conf", 1.0) < VERIFY_CONF_THRESHOLD
             or any(c in item["text"] for c in PUNCT_CHARS))
    ]
    print(f"Pass 4: Checking {len(punct_candidates)} cells for missing punctuation "
          f"(skipped {len(items) - len(punct_candidates)} high-conf/no-punct cells)...")
    def _punct_task(args):
        return check_missing_punct(img, args[1])
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(_punct_task, arg): arg for arg in punct_candidates}
        done = 0
        for future in as_completed(futures):
            done += 1
            if done % 50 == 0:
                print(f"  ... {done}/{len(punct_candidates)}")
            try:
                future.result()
            except Exception as e:
                print(f"  [warn] punct check error: {e}")
    # ── Pass 5: Pattern-based OCR error detection ─────────────────────────────
    print("Pass 5: Pattern-based OCR error detection...")
    pattern_flagged = 0
    for item in items:
        if item.get("missed"):
            item["pattern_error"] = True
            item["pattern_reason"] = "Region completely unread by OCR"
            pattern_flagged += 1
            continue
        has_error, reason = detect_pattern_errors(item["text"])
        if has_error:
            item["pattern_error"] = True
            item["pattern_reason"] = reason
            pattern_flagged += 1
    print(f"  → {pattern_flagged} cells flagged by pattern detection")
    flagged_punct = sum(1 for it in items if it.get("missing_punct"))
    print(f"  → {flagged_punct} cells flagged with missing punctuation")
    return items
# ─────────────────────────────────────────────────────────────────────────────
# GROUPING
# ─────────────────────────────────────────────────────────────────────────────
def group_into_lines(items):
    items = sorted(items, key=lambda x: x["y_center"])
    lines = []
    for item in items:
        placed = False
        for line in lines:
            cy = np.mean([i["y_center"] for i in line])
            if abs(item["y_center"] - cy) < item["height"] * 0.7:
                line.append(item)
                placed = True
                break
        if not placed:
            lines.append([item])
    for line in lines:
        line.sort(key=lambda x: x["x_start"])
    lines.sort(key=lambda ln: np.mean([i["y_center"] for i in ln]))
    return lines
def group_lines_into_records(lines):
    if not lines:
        return []
    records = [[lines[0]]]
    for i in range(1, len(lines)):
        prev_cy = np.mean([it["y_center"] for it in lines[i - 1]])
        curr_cy = np.mean([it["y_center"] for it in lines[i]])
        prev_h  = np.mean([it["height"]   for it in lines[i - 1]])
        if curr_cy - prev_cy > prev_h * 2.0:
            records.append([lines[i]])
        else:
            records[-1].append(lines[i])
    return records
# ─────────────────────────────────────────────────────────────────────────────
# WRITE EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def write_excel(records, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OCR_Output"
    # ┌─────────────────────────────────────────────────────────────────────┐
    # │  RED     = Data NOT properly extracted / character(s) missing       │
    # │  No fill = Data correctly extracted — no issue                      │
    # └─────────────────────────────────────────────────────────────────────┘
    fill_red   = PatternFill("solid", fgColor="FF0000")
    font_white = Font(color="FFFFFF", name="Arial", size=11, bold=True)
    font_black = Font(color="000000", name="Arial", size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    def is_bad_extraction(item):
        """
        Returns (is_bad: bool, reason: str).
        A cell is RED when ANY of these are true:
          • completely unread (???)
          • pattern-detected missing character (parenthesis, x, comma, space)
          • missing punctuation confirmed by re-OCR
          • missing letters confirmed by re-OCR
          • extra garbage / mangled text confirmed by re-OCR
          • very low confidence (< LOW_CONF_THRESHOLD)
        """
        if item.get("missed"):
            return True, "Region completely unread by OCR"
        if item.get("pattern_error"):
            return True, item.get("pattern_reason", "Pattern-based OCR error detected")
        if item.get("missing_punct"):
            reocr = item.get("punct_reocr", "")
            return True, f"Missing punctuation — image shows: {reocr}"
        if item.get("missing_letters"):
            return True, "Incomplete word — letters missing from extraction"
        if item.get("extra_garbage"):
            return True, "Extra garbage characters detected"
        c = item.get("conf", 1.0)
        if c >= 0 and c < LOW_CONF_THRESHOLD:
            return True, f"Very low OCR confidence ({c:.2f}) — likely incorrect"
        return False, None
    row = 1
    total_bad = 0
    for record in records:
        for line in record:
            for col, item in enumerate(line, 1):
                cell = ws.cell(row=row, column=col, value=item.get("text", ""))
                bad, reason = is_bad_extraction(item)
                if bad:
                    cell.fill = fill_red
                    cell.font = font_white
                    total_bad += 1
                    cmt = Comment(f"OCR Issue: {reason}", "OCR Checker")
                    cmt.width  = 270
                    cmt.height = 60
                    cell.comment = cmt
                else:
                    cell.font = font_black
                cell.border    = border
                cell.alignment = Alignment(wrap_text=False, vertical="center")
            row += 1
        row += 1  # blank row between records
    print(f"  → {total_bad} cells highlighted RED (not properly extracted)")
    # ── Legend sheet ──────────────────────────────────────────────────────────
    leg = wb.create_sheet("Legend")
    legend_rows = [
        ("Color",   "Meaning",                                                      None),
        ("RED",     "Data NOT properly extracted — character(s) missing or unread", "FF0000"),
        ("No fill", "Data correctly extracted — no OCR issue detected",             None),
    ]
    for i, (label, meaning, hex_) in enumerate(legend_rows, 1):
        c1 = leg.cell(i, 1, label)
        c2 = leg.cell(i, 2, meaning)
        bold = (i == 1)
        if hex_:
            c1.fill = PatternFill("solid", fgColor=hex_)
            c1.font = Font(color="FFFFFF", name="Arial", size=11, bold=bold)
        else:
            c1.font = Font(name="Arial", size=11, bold=bold)
        c2.font   = Font(name="Arial", size=11, bold=bold)
        c1.border = border
        c2.border = border
    leg.column_dimensions['A'].width = 12
    leg.column_dimensions['B'].width = 65
    # ── Auto-fit column widths on main sheet ──────────────────────────────────
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)
    wb.save(output_path)
    print(f"\nSaved → {output_path}")
# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
def main():
    t0 = time.time()
    print(f"Image  : {IMAGE_PATH}")
    print(f"Output : {OUTPUT_XLSX}\n")
    print("Step 1/4  Preprocessing image...")
    img = preprocess_full(IMAGE_PATH)
    print("Step 2/4  Running OCR pipeline...")
    items = run_ocr(img)
    print("Step 3/4  Grouping into lines and records...")
    lines   = group_into_lines(items)
    records = group_lines_into_records(lines)
    print("Step 4/4  Writing Excel...")
    write_excel(records, OUTPUT_XLSX)
    print(f"\nDone — total time: {time.time() - t0:.1f}s")
if __name__ == "__main__":
    main()