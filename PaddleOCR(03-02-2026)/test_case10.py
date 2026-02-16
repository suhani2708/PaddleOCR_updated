# -*- coding: utf-8 -*-


import os
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter



IMAGE_PATH = 'docs/images/Statista-FormF-RKJJ-NA-1929751_d.jpg'
OUTPUT_FILE = 'my_table.xlsx'

X_TOLERANCE = 60  # Increased for better column grouping (was 40)
Y_TOLERANCE = 25  # Vertical spacing within columns

# ============================================================================


def extract_column_first(image_path, output_excel, x_tol=60, y_tol=25):
    """Extract text column-by-column (X→Y sorting) and save to Excel"""
    
    # print("="*80)
   
    # print("="*80)
    
    # Step 1: Check if image exists
    if not os.path.exists(image_path):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        image_path = os.path.join(script_dir, os.path.basename(image_path))
        
        if not os.path.exists(image_path):
            print(f"\n ERROR: Image not found!")
            print(f"Searched for: {os.path.basename(image_path)}")
            print(f"In folder: {script_dir}")
            print("\n SOLUTIONS:")
            print(f"1. Copy image to script folder")
            print(f"2. Or use full path in IMAGE_PATH")
            return None
    
    # print(f"\n Found image: {os.path.basename(image_path)}")
    
    # Step 2: Initialize OCR
    # print("\n[1/4] Initializing OCR...")
    ocr = PaddleOCR(lang='en', use_gpu=False, show_log=False)
    # print("✓ Done")
    
    # Step 3: Extract text + coordinates
    # print("\n[2/4] Extracting text...")
    result = ocr.ocr(image_path, cls=True)
    
    if not result or not result[0]:
        print("❌ No text detected!")
        return None
    
    # print(f"✓ Found {len(result[0])} text regions")
    
    # Step 4: Prepare boxes with centers
    boxes = []
    for det in result[0]:
        coords = det[0]
        text = det[1][0].strip()
        conf = det[1][1] * 100
        
        y_vals = [p[1] for p in coords]
        x_vals = [p[0] for p in coords]
        
        boxes.append({
            'text': text,
            'conf': conf,
            'x': (min(x_vals) + max(x_vals)) / 2,
            'y': (min(y_vals) + max(y_vals)) / 2,
        })
    

    
    
    # STEP 1: Sort ALL boxes by X (left to right)
    boxes.sort(key=lambda b: b['x'])
    
    # STEP 2: Group into vertical columns using X tolerance
    columns = []
    if boxes:
        current_col = [boxes[0]]
        for box in boxes[1:]:
            avg_x = sum(b['x'] for b in current_col) / len(current_col)
            if abs(box['x'] - avg_x) <= x_tol:
                current_col.append(box)
            else:
                columns.append(current_col)
                current_col = [box]
        if current_col:
            columns.append(current_col)
      
    # STEP 3: Within each column, sort by Y (top to bottom)
    for col in columns:
        col.sort(key=lambda b: b['y'])
    
    # print(f"✓ Detected {len(columns)} vertical columns")
    
    # Warn if too many columns (likely X_TOLERANCE too small)
    if len(columns) > 15:
        print(f"  Warning: {len(columns)} columns detected. Consider increasing X_TOLERANCE if this seems excessive.")
    
    # STEP 4: Build rows from columns (transpose column data into row format)
    max_rows = max(len(col) for col in columns) if columns else 0
    rows = []
    for row_idx in range(max_rows):
        row = []
        for col in columns:
            if row_idx < len(col):
                row.append(col[row_idx])
            else:
                row.append({'text': '', 'conf': 100})  # Empty cell placeholder
        rows.append(row)
    
    # Preview
    # print(f"✓ Organized into {len(rows)} rows across {len(columns)} columns")
    # print("\n Preview (first 3 rows):")
    # for i, row in enumerate(rows[:3], 1):
    #     preview = " | ".join([b['text'][:25] for b in row[:4]])
    #     if len(row) > 4:
    #         preview += f" | ... (+{len(row)-4} cols)"
    #     print(f"  Row {i}: {preview}")
    
    # Step 5: Save to Excel (clean output - no borders/headers)
    # print(f"\n[4/4] Saving to Excel...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Table"
    
    yellow_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
    
    # Write data starting at A1 (no headers, no row numbers)
    for row_idx, row in enumerate(rows):
        for col_idx, box in enumerate(row):
            cell = ws.cell(row_idx + 1, col_idx + 1, box['text'])
            if box['conf'] < 100:
                cell.fill = yellow_fill
    
    # Auto-adjust column widths - PROPER HANDLING FOR 27+ COLUMNS
    for col_idx in range(len(columns)):
        col_letter = get_column_letter(col_idx + 1) 
        ws.column_dimensions[col_letter].width = 28
    
    # Save
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), output_excel)
    wb.save(output_path)
    
    # print(f"✓ Saved to: {output_path}")
    # print(f"\n{'='*80}")
    # print(" SUCCESS - COLUMN-FIRST (X→Y) EXTRACTION!")
    # print("="*80)
    # print(f" File: {output_path}")
    # print(f" Columns detected: {len(columns)}")
    # print(f" Rows extracted: {len(rows)}")
    # print(" Sorting: Left→Right columns → Top→Bottom within each column")
    # print("="*80)
    
    return output_path


# RUN EXTRACTION
if __name__ == "__main__":
    extract_column_first(IMAGE_PATH, OUTPUT_FILE, X_TOLERANCE, Y_TOLERANCE)